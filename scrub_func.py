import openpyxl as xl
from openpyxl.utils import get_column_letter
from tkinter import filedialog

def get_sheets(file):
    wb = xl.load_workbook(file)
    sheet_list = wb.sheetnames()

def choose_sheets(sheets):
    sub_win = tk.Toplevel()
    sub_win.title('Sheet Selection')
    for sheet in sheets:
        pass
    select = tk.Button(text='Okay', command= sub_win.destroy() )

def scrub_file(orig,sheetname,column_list,new_name,save_loc):

    #org = xl.load_workbook(orig)
    sheet = org[sheetname]

    # need to get Equipment Description, Sample Point,
    # Collection Date, Comments and everything to the right of Comments
    collect = column_list

    #Functions
    def getColumn(old_col, new_col, ws, new_ws):
        for row in range(1,ws.max_row+1):
            #print(row)
            new_ws[get_column_letter(new_col)+str(row)] = ws[get_column_letter(old_col)+str(row)].value


    #Create New workbook
    print('Creating new workbook...')
    new_excel = xl.Workbook()
    #Create New sheet
    newsheet = new_excel.active
    newsheet.title = sheetname+'(scrubbed)'


    print('Finding and scrubbing excel info...')
    #equip = sheet['H']
    #sample_point = sheet['I']
    #collection_date = sheet['N']
    #comments = sheet['O']
    #for x in range(1, sheet.max_column):
        #print(x)
        #print(sheet[get_column_letter(x)+str(1)].value, 'Column:', get_column_letter(x))
        #    pass
    newColumn = 1
    for top in range(1, sheet.max_column):
        title = sheet[get_column_letter(top)+str(1)].value
        #print(title)
        if title in collect:
            getColumn(top, newColumn, sheet, newsheet)
            newColumn+=1

    new_excel.save(save_loc+new_name)
    print('Finished!')
