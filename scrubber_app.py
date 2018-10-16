from tkinter import filedialog
import tkinter as tk
import openpyxl as xl
from openpyxl.utils import get_column_letter
import os

def open_and_save():
    new_excel.save(new_entry.get())
    os.system('xdg-open "'+ new_entry.get()+'"')

def saveIt():
    new_excel.save(new_entry.get())


def get_file():
    filename =  filedialog.askopenfilename(initialdir = "/home/creedg/Desktop/",title = "Select file",filetypes = (("excel files","*.xlsx"),("all files","*.*")))
    file_entry.delete(0, tk.END)
    file_entry.insert(0,str(filename))


def save_filename():
    new_filename = filedialog.asksaveasfilename(initialdir='/home/creedg/', title='Select file', filetypes = (("excel files", "*.xlsx"),("all files", "*.*")))
    new_entry.delete(0, tk.END)
    new_entry.insert(0,new_filename)

def scrub_file(orig):
    progress_label = tk.Label(text='Stage of Scrub')
    progress_label.grid(column=0,row=5)

    progress_entry = tk.Entry()
    progress_entry.grid(column=1,row=5)
    progress_entry.insert(0, 'Starting Scrub!')

    org = xl.load_workbook(orig)
    sheet = org['EA_CR_SI']

    # need to get Equipment Description, Sample Point,
    # Collection Date, Comments and everything to the right of Comments
    collect = ['Equipment Description', "Sample Point", "Collection Date",
    'Comments', 'Barium', 'Boron', 'Calcium', 'Chemical', 'Chemical Residual',
    'Iron', 'Magnesium', 'Manganese', 'Measured Sodium', 'Phosphorus', 'Potassium',
    'Strontium', 'Water Clarity', 'Zinc']

    #Functions
    def getColumn(old_col, new_col, ws, new_ws):
        for row in range(1,ws.max_row+1):
            #print(row)
            new_ws[get_column_letter(new_col)+str(row)] = ws[get_column_letter(old_col)+str(row)].value

    #Create New sheet
    newsheet = new_excel.active
    newsheet.title = 'EA_CR_SI(scrubbed)'


    progress_entry.delete(0, tk.END)
    progress_entry.insert(0, 'Finding and scrubbing excel info...')

    newColumn = 1
    for top in range(1, sheet.max_column):
        title = sheet[get_column_letter(top)+str(1)].value
        #print(title)
        if title in collect:
            getColumn(top, newColumn, sheet, newsheet)
            newColumn+=1

    progress_entry.delete(0, tk.END)
    progress_entry.insert(0, 'Finished!')

    save_button = tk.Button(text='Save',command = saveIt)
    save_button.grid(column=0, row=6)

    open_save = tk.Button(text='Save and Open',command=open_and_save)
    open_save.grid(column=1, row=6)






def check_if_valid():
    old_file = file_entry.get()
    #print (old_file)
    scrub_file(old_file)

#create window
root = tk.Tk()

#create window specs
root.geometry('500x400')
root.title('Excel Scrubber V 0.0.2')

# old file label
file_label = tk.Label(text='Original File')
file_label.grid(column=0, row=1)

# New file label
new_label = tk.Label(text='New File')
new_label.grid(column=0, row=2)
# save location label
#loc_label = tk.Label(text='Save Location')
#loc_label.grid(column=0, row=3)

# Create Entry fields
file_entry = tk.Entry()
file_entry.grid(column=1,row=1)

new_entry = tk.Entry()
new_entry.grid(column=1,row=2)

#loc_entry = tk.Entry()
#loc_entry.grid(column=1,row=3)

#Create Buttons
file_button = tk.Button(text='Browse', command=get_file)
file_button.grid(column=2,row=1)

newfile_button = tk.Button(text='Save as..', command=save_filename)
newfile_button.grid(column=2,row=2)

scrub_button = tk.Button(text='SCRUB IT!', command=check_if_valid)
scrub_button.grid(column=1,row=3)

#Create workbook so it's global
new_excel = xl.Workbook()


root.mainloop()
